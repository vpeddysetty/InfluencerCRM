import csv
import io
import json
import os
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, UploadFile, File, HTTPException
from pydantic import BaseModel
from dotenv import load_dotenv
from agent_service.mapping_service import MetadataMapper
from agent_service.langgraph_workflow import build_graph
from agent_service.llm_service import OpenAIAdvisor, blend_confidence
from agent_service.retrieval_service import MappingExampleRetriever

load_dotenv()

app = FastAPI(title="Influencer CRM Metadata Mapping Agent")
mapper = MetadataMapper()
workflow = build_graph(mapper)
advisor = OpenAIAdvisor()
retriever = MappingExampleRetriever()
REVIEW_THRESHOLD = float(os.getenv("REVIEW_THRESHOLD", "0.7"))


class MappingRequest(BaseModel):
    spreadsheet_columns: List[str]


class ReviewRecommendation(BaseModel):
    spreadsheet_column: str
    target_entity: str
    target_attribute: str
    confidence: float
    recommendation_type: str = "mapped"
    notes: str = ""
    source: Optional[str] = None


class MappingReviewRequest(BaseModel):
    spreadsheet_columns: List[str]
    recommendations: List[ReviewRecommendation]
    approved: bool = True
    approved_by: str = "system"
    template_name: Optional[str] = None
    source_tab_names: List[str] = []
    sample_values_json: Dict[str, Any] = {}
    quality_score: Optional[float] = None


class MappingApproveRequest(BaseModel):
    spreadsheet_columns: List[str]
    recommendations: List[ReviewRecommendation]
    approved_by: str = "system"
    template_name: Optional[str] = None
    source_tab_names: List[str] = []
    sample_values_json: Dict[str, Any] = {}
    quality_score: Optional[float] = None


def extract_columns_from_upload(file_name: str, contents: bytes) -> List[str]:
    if file_name.lower().endswith(".csv"):
        rows = list(csv.reader(io.StringIO(contents.decode("utf-8-sig"))))
        if not rows:
            return []
        return [value for value in rows[0] if value]

    if file_name.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - depends on installed package
            raise HTTPException(status_code=500, detail="openpyxl is required for .xlsx uploads") from exc

        workbook = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        worksheet = workbook.active
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        workbook.close()
        return [value for value in first_row if value]

    if file_name.lower().endswith(".xls"):
        raise HTTPException(status_code=400, detail=".xls files are not supported; please convert to .xlsx")

    raise HTTPException(status_code=400, detail="Only CSV/XLSX/XLS files are supported")


class ContentDraftRequest(BaseModel):
    kind: str = "brief"          # "brief" | "landing"
    campaign_name: str = ""
    brand_name: str = ""
    product: str = ""
    audience: str = ""
    tone: str = "friendly"
    creator_name: str = ""       # for landing / per-creator variants


class BriefDraft(BaseModel):
    """
    What a 'brief' draft must contain before it may reach the browser.

    <p>The UI writes these five strings straight into form fields and joins `hashtags` with a
    space. Nothing downstream re-checks them, so this model is the only thing standing between a
    malformed model response and a brand editing `42` in the summary box.
    """

    summary: str = ""
    goals: str = ""
    dos: str = ""
    donts: str = ""
    talkingPoints: str = ""
    hashtags: List[str] = []


class LandingDraft(BaseModel):
    """
    What a 'landing' draft must contain.

    <p>Copy here may carry {{creator.name}} / {{coupon.code}} / {{discount}} tokens, which the
    renderer substitutes later. They are ordinary text at this stage.
    """

    hero: str = ""
    body: str = ""
    cta: str = ""


# Each kind's validator, and the fields the UI actually reads for it.
_DRAFT_MODELS = {"brief": BriefDraft, "landing": LandingDraft}


def _coerce_draft(kind: str, raw: Any) -> Optional[Dict[str, Any]]:
    """
    Coerce a raw model response into the shape for `kind`, or None if it cannot be salvaged.

    <p><b>Coerce before validating, not instead of it.</b> A model that returns a number where a
    string belongs, or one hashtag as a bare string rather than a list, is *nearly* right — and
    rejecting the whole draft over it throws away a usable response and sends the user to the
    heuristic fallback for no reason. Anything that survives coercion is then validated, and
    anything that does not is rejected rather than guessed at.

    <p>Returns None rather than raising. The caller's contract is already "None means fall back
    to the heuristic draft", and a bad LLM response is an expected event here, not an error.
    """
    if not isinstance(raw, dict):
        return None

    model = _DRAFT_MODELS.get(kind)
    if model is None:
        return None

    fields = model.model_fields
    coerced: Dict[str, Any] = {}
    for name, field in fields.items():
        value = raw.get(name)
        if value is None:
            continue
        if name == "hashtags":
            # A single tag may arrive as a bare string. Splitting it is what the UI would have
            # done anyway; the alternative is rendering the characters of the word as tags.
            if isinstance(value, str):
                value = [part for part in value.replace(",", " ").split() if part]
            elif isinstance(value, list):
                # `is not None` before str(): str(None) is "None", a truthy string that would
                # sail through the strip() check and render as a literal #None tag.
                value = [str(part) for part in value if part is not None and str(part).strip()]
            else:
                continue
        elif not isinstance(value, str):
            # Numbers and booleans stringify cleanly and are almost always a formatting slip.
            # Lists and dicts are a structural misunderstanding — dropping the field lets the
            # model's default fill in rather than writing "['a', 'b']" into a form field.
            if isinstance(value, (int, float, bool)):
                value = str(value)
            else:
                continue
        coerced[name] = value

    try:
        validated = model(**coerced)
    except Exception:
        return None

    # A draft where every field coerced away is not a draft. Returning it would replace the
    # heuristic fallback with a form full of empty strings, which reads as a broken feature
    # rather than a degraded one.
    data = validated.model_dump()
    if not any(_is_present(value) for value in data.values()):
        return None

    return data


def _is_present(value: Any) -> bool:
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, list):
        return bool(value)
    return value is not None


def _heuristic_draft(req: "ContentDraftRequest") -> Dict[str, Any]:
    """Deterministic fallback used when the LLM is unavailable (no/invalid key)."""
    product = req.product or "our product"
    brand = req.brand_name or "the brand"
    who = req.creator_name or "your favorite creator"
    if req.kind == "landing":
        return {
            "kind": "landing",
            "source": "heuristic",
            "hero": f"{{{{creator.name}}}} loves {product} — grab {{{{discount}}}}",
            "body": f"{who} partnered with {brand} to bring you {product}. Use the code below to save.",
            "cta": "Shop now",
        }
    return {
        "kind": "brief",
        "source": "heuristic",
        "summary": f"{brand} campaign '{req.campaign_name or 'Untitled'}' featuring {product}.",
        "goals": "Drive awareness and coupon-attributed sales through creator content.",
        "dos": f"Show {product} in authentic use; mention the discount code; tag {brand}.",
        "donts": "No unsupported claims; no competitor mentions; don't omit the paid disclosure.",
        "talkingPoints": f"Key benefits of {product}; why it fits the {req.audience or 'target'} audience.",
        "hashtags": ["#ad", "#partner"],
    }


def _llm_draft(req: "ContentDraftRequest") -> Optional[Dict[str, Any]]:
    if not advisor.is_available():
        return None
    ask = (
        "You write influencer-marketing content. Return ONLY strict JSON. "
        f"kind={req.kind}. For kind='brief' return keys: summary, goals, dos, donts, talkingPoints, hashtags (array). "
        "For kind='landing' return keys: hero, body, cta. "
        "In landing copy you MAY use tokens {{creator.name}}, {{coupon.code}}, {{discount}}."
    )
    payload = {
        "campaign_name": req.campaign_name, "brand_name": req.brand_name,
        "product": req.product, "audience": req.audience, "tone": req.tone,
        "creator_name": req.creator_name, "kind": req.kind,
    }
    try:
        response = advisor.client.responses.create(
            model=advisor.model,
            input=[
                {"role": "system", "content": ask},
                {"role": "user", "content": json.dumps(payload)},
            ],
            temperature=0.6,
        )
        content = getattr(response, "output_text", None)
        if not content and hasattr(response, "choices") and response.choices:
            content = response.choices[0].message.content
        if not content:
            return None
        parsed = json.loads(content)
        # Was `isinstance(parsed, dict)` and nothing more, so any dict reached the browser --
        # including one with a number in `summary` or no `dos` key at all. The UI writes these
        # straight into form fields, so a malformed response became something a brand had to
        # notice and delete by hand.
        drafted = _coerce_draft(req.kind, parsed)
        if drafted is None:
            return None
        drafted["kind"] = req.kind
        drafted["source"] = "llm"
        return drafted
    except Exception:
        return None


@app.post("/content/draft")
def content_draft(request: ContentDraftRequest) -> Dict[str, Any]:
    if request.kind not in ("brief", "landing"):
        raise HTTPException(status_code=400, detail="kind must be 'brief' or 'landing'")
    drafted = _llm_draft(request)
    if drafted is None:
        drafted = _heuristic_draft(request)
    return {"status": "ok", "draft": drafted}


class CreatorClassifyRequest(BaseModel):
    """
    Classify a creator from what a platform already told us.

    Note what is NOT here: follower_count, engagement_rate, or any other metric to be
    produced. Metrics are READ from platform APIs (roadmap Phase C decision #4); the model
    only ever labels. Asking a model for a follower count yields a confident, plausible,
    wrong number that a brand would then spend against.
    """
    handle: str = ""
    platform: str = ""
    display_name: str = ""
    recent_captions: str = ""
    # Ask for the deterministic classifier even when a model is configured (roadmap OP-25).
    #
    # Set by the BFF for submissions from the PUBLIC landing-page form, which is reachable
    # without authentication and therefore by people who are not customers. Past a per-page
    # hourly ceiling the caller keeps the lead and takes a keyword niche rather than buying an
    # unbounded number of model calls. It is a request, not a promise of quality -- the result
    # is stamped `heuristic` exactly as any other fallback is, so nothing downstream has to
    # guess which classifier ran.
    prefer_heuristic: bool = False
    # Passed for context only — so "fitness creator, mid-tier audience" is available to the
    # classifier. It is never echoed back as an output.
    follower_count: Optional[int] = None


# The niches a creator may be sorted into. A closed set, because a brand's vetting rules are
# written against these values -- a free-text niche would make "niche not in allowed list"
# unenforceable.
CREATOR_NICHES = [
    "beauty", "fashion", "fitness", "food", "gaming", "lifestyle",
    "parenting", "pets", "sports", "tech", "travel", "finance", "other",
]

# Brand-safety flags. Advisory input to a human decision -- never an automatic approval.
RISK_FLAGS = ["adult", "alcohol", "gambling", "politics", "controversy", "tobacco"]

# Substrings that raise each flag in the deterministic fallback. Deliberately narrow: a
# false positive sends a legitimate creator to a review queue, which is recoverable, while
# a false negative is what a brand would actually complain about.
_RISK_MARKERS = {
    "gambling": ["casino", "bet ", "betting", "poker", "deposit bonus", "odds", "wager"],
    "alcohol": ["drinks", "cocktail", "wine", "beer", "whisky", "vodka"],
    "adult": ["18+", "nsfw", "onlyfans", "adult only"],
    "tobacco": ["vape", "vaping", "cigar", "smoking", "nicotine"],
    "politics": ["election", "vote for", "campaign trail", "senator", "parliament"],
}

_NICHE_MARKERS = {
    "fitness": ["workout", "gym", "protein", "training", "reps", "trainers"],
    "beauty": ["skincare", "serum", "foundation", "makeup", "glow", "routine"],
    "food": ["recipe", "kitchen", "brunch", "baking", "restaurant"],
    "gaming": ["stream", "gameplay", "console", "loadout", "speedrun"],
    "travel": ["flight", "hotel", "itinerary", "packing", "destination"],
    "tech": ["gadget", "unboxing", "benchmark", "laptop", "firmware"],
    "fashion": ["outfit", "haul", "wardrobe", "styling", "lookbook"],
    "parenting": ["toddler", "newborn", "nursery", "mum life", "parenting"],
    "pets": ["puppy", "kitten", "rescue dog", "vet visit"],
    "finance": ["portfolio", "investing", "savings", "budget", "etf"],
}


def _heuristic_classify(req: "CreatorClassifyRequest") -> Dict[str, Any]:
    """
    Deterministic fallback when no LLM is configured.

    Keyword matching, which is genuinely weaker than a model at reading intent -- so it is
    labelled `heuristic` in the response and stored as such. A brand seeing a niche can then
    tell whether a model or a substring match produced it.
    """
    text = f"{req.handle} {req.display_name} {req.recent_captions}".lower()

    risk_flags = sorted({
        flag for flag, markers in _RISK_MARKERS.items()
        if any(marker in text for marker in markers)
    })

    niche = "other"
    best = 0
    for candidate, markers in _NICHE_MARKERS.items():
        hits = sum(1 for marker in markers if marker in text)
        if hits > best:
            best, niche = hits, candidate

    themes = []
    for candidate, markers in _NICHE_MARKERS.items():
        if any(marker in text for marker in markers):
            themes.append(candidate)

    return {
        "source": "heuristic",
        "niche": niche,
        "content_themes": sorted(themes) or ["general"],
        "risk_flags": risk_flags,
        "summary": f"Keyword classification of @{req.handle or 'unknown'} as {niche}.",
    }


def _llm_classify(req: "CreatorClassifyRequest") -> Optional[Dict[str, Any]]:
    if not advisor.is_available():
        return None
    ask = (
        "You classify social media creators for brand partnerships. Return ONLY strict JSON "
        f"with keys: niche (one of {CREATOR_NICHES}), content_themes (array of short strings), "
        f"risk_flags (array, subset of {RISK_FLAGS}, empty when nothing applies), "
        "summary (one sentence). "
        "Classify ONLY from the text provided. Do NOT estimate follower counts, engagement, "
        "reach or any other metric - those are measured elsewhere and inventing them is worse "
        "than omitting them. Raise a risk flag only on clear evidence in the captions."
    )
    payload = {
        "handle": req.handle,
        "platform": req.platform,
        "display_name": req.display_name,
        "recent_captions": req.recent_captions,
        "follower_count": req.follower_count,
    }
    try:
        response = advisor.client.responses.create(
            model=advisor.model,
            input=[
                {"role": "system", "content": ask},
                {"role": "user", "content": json.dumps(payload)},
            ],
            # Lower than the drafting endpoint: classification wants the same answer every
            # time for the same input, not variety.
            temperature=0.1,
        )
        content = getattr(response, "output_text", None)
        if not content and hasattr(response, "choices") and response.choices:
            content = response.choices[0].message.content
        if not content:
            return None
        parsed = json.loads(content)
        if not isinstance(parsed, dict):
            return None

        # Constrain the model to the closed vocabularies. A model returning "beauty & skincare"
        # would silently break any vetting rule written against "beauty".
        niche = str(parsed.get("niche", "other")).strip().lower()
        if niche not in CREATOR_NICHES:
            niche = "other"
        flags = [f for f in parsed.get("risk_flags", []) if f in RISK_FLAGS]
        themes = [str(t)[:40] for t in parsed.get("content_themes", [])][:10]

        return {
            "source": "llm",
            "niche": niche,
            "content_themes": themes or ["general"],
            "risk_flags": sorted(set(flags)),
            "summary": str(parsed.get("summary", ""))[:400],
        }
    except Exception:
        return None


@app.post("/creators/classify")
def creators_classify(request: CreatorClassifyRequest) -> Dict[str, Any]:
    """
    Label a creator. Never produces a metric — see CreatorClassifyRequest.

    Always returns a classification: the heuristic fallback means a signup is never blocked
    because an API key is missing or the model call failed (roadmap C.6).
    """
    if not request.handle and not request.recent_captions:
        raise HTTPException(status_code=400, detail="handle or recent_captions is required")
    # `prefer_heuristic` short-circuits BEFORE the model call rather than discarding its answer
    # afterwards -- the whole point is not to spend the money, so the order matters.
    if request.prefer_heuristic:
        return {"classification": _heuristic_classify(request)}
    result = _llm_classify(request) or _heuristic_classify(request)
    return {"classification": result}


@app.get("/health")
def health() -> Dict[str, Any]:
    """
    Liveness, plus whether the model is actually reachable (roadmap OP-27).

    `status` stays "ok" whenever the process is serving, because that is what the container
    healthcheck and the ASG read -- degrading it on a missing key would take the platform down
    over a feature that is designed to survive without one.

    `llm` is the part worth having. Every LLM path here falls back to a deterministic matcher and
    returns a result, so a missing key produces no error, no 500 and no log line at request time --
    creator vetting silently ran on substring matching in production for three weeks and was found
    only by someone measuring `source` on stored rows. An unavailable model is a real operational
    state, and this is the cheapest place to make it observable rather than inferable.
    """
    return {"status": "ok", "llm": "available" if advisor.is_available() else "unavailable"}


@app.get("/mappings/examples")
def list_mapping_examples(
    limit: int = 20,
    active_only: bool = True,
    template_name: Optional[str] = None,
) -> Dict[str, Any]:
    response = retriever.list_examples(limit=limit, active_only=active_only, template_name=template_name)
    if not response.get("ok"):
        return {
            "status": "error",
            "reason": response.get("reason", "unknown"),
            "error": response.get("error"),
            "items": [],
        }

    return {
        "status": "ok",
        "count": response.get("count", 0),
        "items": response.get("items", []),
    }


@app.post("/mappings/review")
def review_mapping(request: MappingReviewRequest) -> Dict[str, Any]:
    if not request.spreadsheet_columns:
        raise HTTPException(status_code=400, detail="spreadsheet_columns cannot be empty")

    persistence = retriever.save_review_decision(
        spreadsheet_columns=request.spreadsheet_columns,
        recommendations=[item.model_dump() for item in request.recommendations],
        approved=request.approved,
        approved_by=request.approved_by,
        template_name=request.template_name,
        source_tab_names=request.source_tab_names,
        sample_values_json=request.sample_values_json,
        quality_score=request.quality_score,
    )

    if not persistence.get("saved"):
        return {
            "status": "error",
            "decision": "approved" if request.approved else "rejected",
            "persistence": persistence,
        }

    return {
        "status": "ok",
        "decision": "approved" if request.approved else "rejected",
        "persistence": persistence,
    }


@app.post("/mappings/approve")
def approve_mapping(request: MappingApproveRequest) -> Dict[str, Any]:
    review_request = MappingReviewRequest(
        spreadsheet_columns=request.spreadsheet_columns,
        recommendations=request.recommendations,
        approved=True,
        approved_by=request.approved_by,
        template_name=request.template_name,
        source_tab_names=request.source_tab_names,
        sample_values_json=request.sample_values_json,
        quality_score=request.quality_score,
    )
    return review_mapping(review_request)


@app.post("/map-columns")
def map_columns(request: MappingRequest) -> Dict[str, Any]:
    if not request.spreadsheet_columns:
        raise HTTPException(status_code=400, detail="spreadsheet_columns cannot be empty")

    state = workflow.invoke({"spreadsheet_columns": request.spreadsheet_columns})
    result = state.get("recommendations", {})

    llm_available = advisor.is_available()
    retrieval_available = retriever.is_available()
    llm_enhanced = False
    fallback_used = False
    review_candidates = []
    review_trace = []
    retrieved_examples = retriever.retrieve_examples(request.spreadsheet_columns) if retrieval_available else []

    if llm_available:
        enriched = advisor.recommend(
            request.spreadsheet_columns,
            result.get("metadata_catalog", {}),
            retrieved_examples=retrieved_examples,
        )
        if enriched:
            if "recommendations" in enriched:
                heuristic_recommendations = result.get("recommendations", [])
                llm_recommendations = enriched.get("recommendations", [])
                merged = []
                by_column = {item.get("spreadsheet_column"): item for item in llm_recommendations if isinstance(item, dict) and item.get("spreadsheet_column")}
                for item in heuristic_recommendations:
                    if isinstance(item, dict):
                        column_name = item.get("spreadsheet_column")
                        llm_item = by_column.get(column_name)
                        if llm_item and "confidence" in llm_item:
                            item = dict(item)
                            item["confidence"] = blend_confidence(item.get("confidence", 0.0), llm_item.get("confidence"))
                            item["recommendation_type"] = "mapped"
                            item["notes"] = f"{item.get('notes', '')} LLM confidence blended with heuristic score.".strip()
                            item["source"] = "llm_enhanced"
                            # Adopt the LLM's target when the heuristic only reached a
                            # custom_attributes fallback and the LLM points at a real
                            # attribute that exists in the metadata catalog.
                            if item.get("target_attribute") == "custom_attributes":
                                llm_entity = llm_item.get("target_entity")
                                llm_attribute = llm_item.get("target_attribute")
                                catalog_attrs = mapper.metadata_catalog.get(llm_entity, [])
                                if llm_attribute and llm_attribute != "custom_attributes" and llm_attribute in catalog_attrs:
                                    item["target_entity"] = llm_entity
                                    item["target_attribute"] = llm_attribute
                                    item["notes"] = f"{item.get('notes', '')} Adopted LLM target attribute over custom-attribute fallback.".strip()
                        else:
                            item = dict(item)
                            item["source"] = "heuristic"
                        if item.get("confidence", 0.0) < REVIEW_THRESHOLD:
                            review_candidates.append(column_name)
                            review_trace.append({
                                "spreadsheet_column": column_name,
                                "reason": "low_confidence",
                                "confidence": item.get("confidence", 0.0),
                                "target_attribute": item.get("target_attribute"),
                            })
                        merged.append(item)
                result["recommendations"] = merged
            if "custom_fields" in enriched:
                result["custom_fields"] = enriched["custom_fields"]
            llm_enhanced = True
        else:
            llm_enhanced = False
    else:
        llm_enhanced = False

    if not result.get("recommendations"):
        fallback_recommendations = state.get("recommendations", {}).get("recommendations", [])
        result["recommendations"] = [
            {**item, "source": "fallback"} if isinstance(item, dict) else item
            for item in fallback_recommendations
        ]
        fallback_used = True
    if not result.get("custom_fields"):
        result["custom_fields"] = state.get("recommendations", {}).get("custom_fields", [])

    result["debug"] = {
        "llm_available": llm_available,
        "retrieval_available": retrieval_available,
        "retrieved_examples_count": len(retrieved_examples),
        "llm_enhanced": llm_enhanced,
        "fallback_used": fallback_used,
        "recommendation_count": len(result.get("recommendations", [])),
        "review_candidates": review_candidates,
        "review_trace": review_trace,
    }

    return result


@app.post("/map-upload")
async def map_upload(file: UploadFile = File(...)) -> Dict[str, Any]:
    if not file.filename or not file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only CSV/XLSX/XLS files are supported")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    columns = extract_columns_from_upload(file.filename, contents)
    if not columns:
        raise HTTPException(status_code=400, detail="No columns were found in the uploaded file")
    return map_columns(MappingRequest(spreadsheet_columns=columns))
